import os
import sys
from openpyxl import load_workbook

def parse_report(file_path):
    """Parses a CivicBin test report spreadsheet and compiles test case category details."""
    if not os.path.exists(file_path):
        print(f"[-] Error: Report file {file_path} not found.")
        return None
    
    wb = load_workbook(file_path, data_only=True, read_only=True)
    sheet = wb.active
    
    total = 0
    passed = 0
    failed = 0
    category_counts = {}
    
    first_row = True
    for row in sheet.iter_rows(values_only=True):
        if first_row:
            # Skip headers row
            first_row = False
            continue
        
        if len(row) < 6 or not row[0]:
            continue
            
        category = row[2]
        status = row[4]
        
        total += 1
        if status == "PASS":
            passed += 1
        else:
            failed += 1
            
        # Standardize platform-specific categories for side-by-side comparison
        display_category = str(category)
        if "Specific Testing" in display_category:
            display_category = "Platform-Specific Testing"
            
        category_counts[display_category] = category_counts.get(display_category, 0) + 1
        
    return {
        "total": total,
        "passed": passed,
        "failed": failed,
        "categories": category_counts
    }

def main():
    import sys
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
        
    selenium_report = "reports/CivicBin_Web_Selenium_Report.xlsx"
    appium_report = "reports/CivicBin_Mobile_Appium_Report.xlsx"
    
    print("[+] Parsing E2E reports...")
    sel_data = parse_report(selenium_report)
    app_data = parse_report(appium_report)
    
    if not sel_data or not app_data:
        print("[-] Error: Could not load report spreadsheets.")
        sys.exit(1)
        
    total_cases = sel_data["total"] + app_data["total"]
    total_passed = sel_data["passed"] + app_data["passed"]
    total_failed = sel_data["failed"] + app_data["failed"]
    pass_rate = (total_passed / total_cases * 100) if total_cases > 0 else 0
    
    # Collect and sort all unique category names
    categories = sorted(list(set(list(sel_data["categories"].keys()) + list(app_data["categories"].keys()))))
    
    # Construct Markdown text matching the user's requested layout
    md = []
    md.append("# CivicBin CI/CD Pipeline — Test Results\n")
    
    # Table 1: Summary Table
    md.append("| Test Suite | Platform | Test Cases | Passed | Failed | Pass Rate |")
    md.append("| :--- | :--- | :---: | :---: | :---: | :---: |")
    
    sel_rate = (sel_data["passed"] / sel_data["total"] * 100) if sel_data["total"] > 0 else 0
    app_rate = (app_data["passed"] / app_data["total"] * 100) if app_data["total"] > 0 else 0
    
    sel_status_icon = "✅" if sel_data["failed"] == 0 else "❌"
    app_status_icon = "✅" if app_data["failed"] == 0 else "❌"
    
    md.append(f"| Selenium E2E Web + API Tests | Web (Chrome) | {sel_data['total']} | {sel_data['passed']} | {sel_data['failed']} | {sel_status_icon} {sel_rate:.0f}% |")
    md.append(f"| Appium Mobile Tests Dry-Run | Android | {app_data['total']} | {app_data['passed']} | {app_data['failed']} | {app_status_icon} {app_rate:.0f}% |")
    md.append("\n")
    
    # Overall summary line
    md.append(f"## Overall: {total_passed} / {total_cases} test cases passed — {pass_rate:.0f}% Pass Rate 🎉\n")
    
    # Table 2: Category Breakdown Table
    md.append("| Category | Selenium Tests | Appium Tests |")
    md.append("| :--- | :---: | :---: |")
    
    for cat in categories:
        sel_count = sel_data["categories"].get(cat, 0)
        app_count = app_data["categories"].get(cat, 0)
        md.append(f"| {cat} | {sel_count} | {app_count} |")
        
    md.append(f"| **Total** | **{sel_data['total']}** | **{app_data['total']}** |")
    md.append("\n")
    md.append("Job summary generated at run-time\n")
    
    markdown_content = "\n".join(md)
    
    # Output markdown to terminal logs
    print(markdown_content)
    
    # Write to GITHUB_STEP_SUMMARY
    github_summary = os.environ.get("GITHUB_STEP_SUMMARY")
    if github_summary:
        with open(github_summary, "w", encoding="utf-8") as f:
            f.write(markdown_content)
        print("[+] Written summary to GITHUB_STEP_SUMMARY successfully.")
    else:
        print("[-] GITHUB_STEP_SUMMARY environment variable not set (local run).")

if __name__ == "__main__":
    main()
